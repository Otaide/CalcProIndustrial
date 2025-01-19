from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import tkinter.filedialog as filedialog
import webbrowser
from historico import HistoricoFrame
from pathlib import Path
import os
from subprocess import Popen
import sys

# Função para carregar dados do CSV
def load_data_from_csv(filepath):
    formula_data = {}
    try:
        with open(filepath, mode='r', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            for row in reader:
                try:
                    formula_id = int(row['Fórmula'].split()[0])
                    description = row['Descrição']
                    fixed_amount = float(row['Kg'])
                    type_ = row.get('Tipo', 'N/A')
                    obs = row.get('Observação', 'N/A')

                    if formula_id not in formula_data:
                        formula_data[formula_id] = []

                    formula_data[formula_id].append({
                        'description': description,
                        'fixed_amount': fixed_amount,
                        'type': type_,
                        'obs': obs
                    })
                    
                except ValueError:
                    print(f"Valor inválido para fórmula: {row['Fórmula']}")
                except KeyError as e:
                    print(f"Coluna ausente no CSV: {e}")
                    raise
    except FileNotFoundError:
        print(f"Arquivo não encontrado: {filepath}")
    except Exception as e:
        print(f"Erro ao ler o arquivo CSV: {e}")
    return formula_data

# Carregar dados do CSV
formula_data = load_data_from_csv('bd.csv')

class PlaceholderEntry(ttk.Entry):
    def __init__(self, master=None, placeholder="", **kwargs):
        super().__init__(master, **kwargs)
        self.placeholder = placeholder
        self.bind("<FocusIn>", self.remove_placeholder)
        self.bind("<FocusOut>", self.add_placeholder)
        self.add_placeholder()

    def add_placeholder(self, event=None):
        if not self.get():
            self.insert(0, self.placeholder)
            self.configure(foreground='grey')

    def remove_placeholder(self, event=None):
        if self.get() == self.placeholder:
            self.delete(0, tk.END)
            self.configure(foreground='black')

class CalculatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Calculadora de Fórmulas")
        self.root.state('zoomed')  # Janela maximizada
        self.style = ttk.Style()
        self.configure_styles()
        self.create_widgets()
        self.original_results = []  # Para armazenar todos os resultados possíveis
        self.calculated_results = []  # Para armazenar os resultados do cálculo atual
        self.filtro = None

        # Bind eventos de scroll
        self.form.bind('<Configure>', lambda e: self.form_canvas.configure(scrollregion=self.form_canvas.bbox("all")))
        self.form_canvas.bind('<Configure>', self.on_canvas_configure)

    def configure_styles(self):
        # Cores do tema
        self.colors = {
            'primary': '#2196F3',      # Azul principal
            'secondary': '#607D8B',    # Azul acinzentado
            'background': '#F5F5F5',   # Cinza muito claro
            'surface': '#FFFFFF',      # Branco
            'accent': '#FF4081',       # Rosa accent
            'text': '#212121',         # Texto quase preto
            'text_secondary': '#757575' # Texto cinza
        }

        # Estilo para o Frame principal
        self.style.configure(
            "TFrame",
            background=self.colors['background']
        )

        # Estilo para frames de seção
        self.style.configure(
            "Section.TFrame",
            background=self.colors['surface'],
            relief="raised",
            borderwidth=1
        )

        # Estilo para toolbar
        self.style.configure(
            "Toolbar.TFrame",
            background=self.colors['primary'],
            relief="flat"
        )

        # Estilo para labels
        self.style.configure(
            "TLabel",
            background=self.colors['background'],
            foreground=self.colors['text'],
            font=("Segoe UI", 11)
        )

        # Estilo para label do desenvolvedor
        self.style.configure(
            "Dev.TLabel",
            background=self.colors['background'],
            foreground=self.colors['text_secondary'],
            font=("Segoe UI", 10, "italic")
        )

        # Estilo para Treeview
        self.style.configure(
            "Treeview",
            background=self.colors['surface'],
            foreground=self.colors['text'],
            fieldbackground=self.colors['surface'],
            font=("Segoe UI", 11)
        )

        # Estilo para cabeçalhos do Treeview
        self.style.configure(
            "Treeview.Heading",
            background=self.colors['primary'],
            foreground=self.colors['surface'],
            font=("Segoe UI", 11, "bold")
        )

        # Estilos para os Labels
        self.style.configure("TLabel", background="#f4f4f4", font=("Segoe UI", 17), foreground="#333333")
        
        # Estilos para os Botões
        self.style.configure("TButton", font=("Segoe UI", 14), padding=6)
        
        # Botão Pesquisar e Buscar
        self.style.configure("Search.TButton", background="#0cccf2", foreground="#e01429", padding=6)
        self.style.configure("Filter.TButton", background="#edda0c", foreground="#010212", padding=6)
        
        # Botão Calcular
        self.style.configure("Calculate.TButton", background="#11f258", foreground="#0eed25", padding=6)
        
        # Estilos para Checkbuttons
        self.style.configure("TCheckbutton", background="#f4f4f4", font=("Segoe UI", 14), foreground="#333333")
        
        # Estilos para Treeview Heading
        self.style.configure("Treeview.Heading", font=("Segoe UI", 16, 'bold'), background="#edde0e", foreground="#0f0101")

        # Estilos para Treeview
        self.style.configure("Treeview", font=("Segoe UI", 13), background="#0f0505", foreground="#faf2f2", fieldbackground="#f9f9f9")
        self.style.map("Treeview",
            background=[('selected', '#e64562')],
            foreground=[('selected', '#f5eded')]
        )

        # Estilo para o botão de histórico
        self.style.configure("History.TButton",
            background="#F1C40F",  # Amarelo
            foreground="#000000",  # Texto preto
            padding=8,
            font=("Segoe UI", 12, "bold"),
            relief="raised",
            borderwidth=2
        )
        
        self.style.map("History.TButton",
            background=[
                ("pressed", "#F39C12"),  # Amarelo mais escuro quando pressionado
                ("active", "#F4D03F"),   # Amarelo mais claro quando hover
                ("disabled", "#BDC3C7")  # Cinza quando desabilitado
            ],
            foreground=[
                ("pressed", "#000000"),  # Mantém texto preto em todos os estados
                ("active", "#000000"),
                ("disabled", "#95A5A6")
            ],
            relief=[
                ("pressed", "sunken"),
                ("!pressed", "raised")
            ]
        )

        # Estilo para o botão de informações
        self.style.configure(
            "Info.TButton",
            background="#e0e0e0",        # Cinza claro
            foreground="#1565c0",        # Azul escuro para o texto
            padding=8,
            font=("Segoe UI", 12, "bold")
        )
        self.style.map("Info.TButton",
            background=[("active", "#d5d5d5")],  # Cinza mais escuro no hover
            foreground=[("active", "#0d47a1")]   # Azul mais escuro no hover
        )

    def adjust_column_width(self):
        """ Ajusta a largura das colunas com base no conteúdo """
        for col in self.results_tree['columns']:
            max_width = len(str(col)) * 10  # Largura mínima baseada no título
            
            # Verificar largura necessária para cada item
            for item in self.results_tree.get_children():
                cell_value = str(self.results_tree.item(item)['values'][self.results_tree['columns'].index(col)])
                cell_width = len(cell_value) * 10
                max_width = max(max_width, cell_width)
            
            # Ajuste fixo baseado em um tamanho de fonte padrão
            final_width = max_width + 20  # +20 para padding
            
            self.results_tree.column(col, width=final_width)

    def create_widgets(self):
        # Container Principal
        self.container = ttk.Frame(self.root, style="TFrame", padding=20)
        self.container.pack(fill=tk.BOTH, expand=True)

        # Frame para informações do desenvolvedor
        dev_frame = ttk.Frame(self.container, style="TFrame")
        dev_frame.pack(fill=tk.X, pady=(0, 10))

        dev_label = ttk.Label(
            dev_frame,
            text="Desenvolvido por Otaide Ferreira - Uso Industrial/Interno",
            style="Dev.TLabel"
        )
        dev_label.pack(side=tk.RIGHT, padx=5)

        # Toolbar com novo estilo
        self.toolbar_frame = ttk.Frame(self.container, style="Toolbar.TFrame")
        self.toolbar_frame.pack(fill=tk.X, pady=(0, 10))

        # Botão Histórico
        self.historico_button = ttk.Button(
            self.toolbar_frame,
            text="Histórico",
            command=self.show_historico,
            style="History.TButton"
        )
        self.historico_button.pack(side=tk.RIGHT, padx=5)

        # Novo Botão Informações
        self.info_button = ttk.Button(
            self.toolbar_frame,
            text="Informações",
            command=self.open_documentation,
            style="Info.TButton"
        )
        self.info_button.pack(side=tk.RIGHT, padx=5)

        # Campo para Nome da Programação
        ttk.Label(
            self.toolbar_frame, 
            text="Nome da Programação:",
            font=("Segoe UI", 12)  # Aumentado tamanho da fonte
        ).pack(side=tk.LEFT, padx=(0, 5))
        
        self.programacao_entry = ttk.Entry(
            self.toolbar_frame, 
            width=50,  # Aumentado de 30 para 50
            font=("Segoe UI", 12)  # Fonte maior para melhor visibilidade
        )
        self.programacao_entry.pack(side=tk.LEFT, padx=5, pady=10)  # Adicionado pady para mais espaço vertical

        # Seção de Fórmulas com novo estilo
        self.form_section = ttk.Frame(
            self.container,
            style="Section.TFrame",
            padding=15
        )
        self.form_section.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Seção de Resultados com novo estilo
        self.results_section = ttk.Frame(
            self.container,
            style="Section.TFrame",
            padding=15
        )
        self.results_section.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Seção de Fórmulas Filtradas
        self.filtered_form_section = ttk.Frame(self.container, padding=15, relief=tk.RAISED, style="TFrame")
        self.filtered_form_section.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.filtered_form_section.pack_forget()  # Inicialmente escondido

        # Barra de Pesquisa para Fórmulas
        self.search_frame = ttk.Frame(self.form_section, padding=10)
        self.search_frame.pack(fill=tk.X, pady=(0, 15))

        ttk.Label(self.search_frame, text="Fórmula por Número:", font=("Segoe UI", 14)).grid(row=0, column=0, sticky=tk.W)
        self.search_entry = PlaceholderEntry(self.search_frame, placeholder="Digite o número", font=("Segoe UI", 14))
        self.search_entry.grid(row=0, column=1, sticky=tk.EW, padx=(5, 0))
        self.search_button = ttk.Button(
            self.search_frame, 
            text="Buscar", 
            command=self.search_formula, 
            style="Search.TButton",
            width=10  # Ajustando largura do botão
        )
        self.search_button.grid(row=0, column=2, padx=(5, 0), pady=(5, 0), ipadx=10, ipady=5)
        
        # Adicionar bind para Enter na pesquisa de fórmulas
        self.search_entry.bind('<Return>', lambda e: self.search_formula())

      # Botão Atualizar Fórmulas para abrir o arquivo HTML
        self.update_formulas_button = ttk.Button(self.search_frame, text="Atualizar Fórmulas", command=self.open_html_file, style="TButton")
        self.update_formulas_button.grid(row=1, column=0, columnspan=1, padx=(10, 10),pady=(5, 2))
        self.clear_checkboxes_button = ttk.Button(self.search_frame, text="Limpar Checkboxes", command=self.clear_checkboxes, style="TButton")
        self.clear_checkboxes_button.grid(row=3, column=2, padx=(5, 0), pady=(10, 0))

        # Botão Calcular posicionado abaixo dos botões de pesquisa
        self.calculate_button = ttk.Button(self.search_frame, text="Calcular", command=self.calculate, style="Calculate.TButton")
        self.calculate_button.grid(row=2, column=0, columnspan=3, pady=(10, 0),sticky="ew")  # Adicionado na linha 2, ocupando 3 colunas

        self.view_all_formulas_button = ttk.Button(self.search_frame, text="Ver Todas as Fórmulas", command=self.show_all_formulas, style="TButton")
        self.view_all_formulas_button.grid(row=3, column=0, columnspan=3, pady=(10, 0))

        # Barra de Pesquisa para Resultados
        self.results_search_frame = ttk.Frame(self.results_section, padding=10)
        self.results_search_frame.pack(fill=tk.X, pady=(0, 15))

        ttk.Label(self.results_search_frame, text="Por Descrição:", font=("Segoe UI", 14)).grid(row=0, column=0, sticky=tk.W)
        self.results_search_entry = PlaceholderEntry(self.results_search_frame, placeholder="Digite a descrição", font=("Segoe UI", 14))
        self.results_search_entry.grid(row=0, column=1, sticky=tk.EW, padx=(5, 0))
        self.results_search_button = ttk.Button(
            self.results_search_frame, 
            text="Pesquisar", 
            command=self.filter_results, 
            style="Filter.TButton",
            width=10  # Ajustando largura do botão
        )
        self.results_search_button.grid(row=0, column=2, padx=(5, 0), ipadx=10, ipady=5)
        
        # Adicionar bind para Enter na pesquisa de resultados
        self.results_search_entry.bind('<Return>', lambda e: self.filter_results())
        # Adicionar bind para ESC para limpar filtro
        self.results_search_entry.bind('<Escape>', lambda e: self.clear_filter())
        self.root.bind('<Escape>', lambda e: self.clear_filter())

        self.clear_filter_button = ttk.Button(self.results_search_frame, text="Limpar Filtro", command=self.clear_filter, style="TButton")
        self.clear_filter_button.grid(row=1, column=0, padx=(5, 0), pady=(5, 0))

        # Frame para Formulário com Scrollbar
        self.form_frame = ttk.Frame(self.form_section, padding=10, relief="flat")
        self.form_frame.pack(fill=tk.BOTH, expand=True)

        # Canvas com scrollbar
        self.form_canvas = tk.Canvas(self.form_frame, bg="#f4f4f4", highlightthickness=0)
        self.form_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Scrollbar vertical
        scrollbar_formula = ttk.Scrollbar(self.form_frame, orient=tk.VERTICAL, command=self.form_canvas.yview)
        scrollbar_formula.pack(side=tk.RIGHT, fill=tk.Y)

        # Configurar o canvas
        self.form_canvas.configure(yscrollcommand=scrollbar_formula.set)
        
        # Frame interno para as fórmulas
        self.form = ttk.Frame(self.form_canvas, padding=10)
        self.form_canvas.create_window((0, 0), window=self.form, anchor="nw", width=self.form_canvas.winfo_width())

        # Bind eventos de scroll
        self.form.bind('<Configure>', lambda e: self.form_canvas.configure(scrollregion=self.form_canvas.bbox("all")))
        self.form_canvas.bind('<Configure>', self.on_canvas_configure)

    # Con# Bind eventos do mouse e teclado para scroll apenas quando o mouse estiver sobre o frame
        self.form_canvas.bind("<Enter>", self.bind_scroll)
        self.form_canvas.bind("<Leave>", self.unbind_scroll)

        self.formulas = {}
        
        self.soma_kg_label = ttk.Label(self.results_search_frame, text="Soma em Kg:")
        self.soma_kg_label.grid(row=2, column=0, padx=(5, 0), pady=(5, 0))

        for formula_id in formula_data.keys():
            formula_frame = ttk.Frame(self.form, style="TFrame", padding=10)
            formula_frame.pack(fill=tk.X, pady=5)

            var = tk.IntVar()
            weight_entry = PlaceholderEntry(formula_frame, placeholder="Digite o peso aqui", font=("Segoe UI", 18))
            weight_entry.config(state=tk.DISABLED)
            checkbox = ttk.Checkbutton(
                formula_frame, 
                text=f"Fórmula {formula_id}", 
                variable=var,
                command=lambda var=var, weight_entry=weight_entry: self.toggle_entry(var, weight_entry),
                style="TCheckbutton"
            )

            checkbox.grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
            weight_entry.grid(row=0, column=1, sticky=tk.EW)

            self.formulas[formula_id] = {
                'checkbox': checkbox,
                'weight_entry': weight_entry,
                'var': var
            }

        # Configurar colunas com cores específicas
        columns_config = {
            "Fórmula": {
                "width":89, 
                "anchor": "center", 
                "stretch": False,  # Alterado para False para manter o tamanho fixo
                "background": "#FF7F50",
                "foreground": "#FFFFFF"
            },
            "Descrição": {
                "width": 180,  # Reduzido ainda mais
                "anchor": "w", 
                "stretch": True,  # Mantém True para ajustar ao conteúdo
                "background": "#4CAF50",
                "foreground": "#FFFFFF"
            },
            "Kg": {
                "width": 80, 
                "anchor": "center", 
                "stretch": False,  # Alterado para False
                "background": "#2196F3",
                "foreground": "#FFFFFF"
            },
            "Tipo": {
                "width": 100, 
                "anchor": "center", 
                "stretch": False,  # Alterado para False
                "background": "#9C27B0",
                "foreground": "#FFFFFF"
            },
            "Observação": {
                "width": 150,
                "anchor": "w", 
                "stretch": True,  # Mantém True para ajustar ao conteúdo
                "background": "#F44336",
                "foreground": "#FFFFFF"
            }
        }

        # Frame para Treeview com Scrollbar
        self.tree_frame = ttk.Frame(self.results_section)
        self.tree_frame.pack(fill=tk.BOTH, expand=True)

        # Criar o Treeview com configuração de altura
        self.results_tree = ttk.Treeview(
            self.tree_frame,
            columns=("Fórmula", "Descrição", "Kg", "Tipo", "Observação"),
            show="headings",
            style="Custom.Treeview",
            selectmode="browse",
            height=20  # Define uma altura fixa em número de linhas
        )

        # Primeiro, configure o estilo base da Treeview
        self.style.configure(
            "Custom.Treeview",
            background="#FFFFFF",
            foreground="#333333",
            fieldbackground="#FFFFFF",
            rowheight=45,
            font=("Segoe UI", 12),
            borderwidth=0,
            relief="flat"
        )
        
        # Depois, configure explicitamente o estilo do cabeçalho
        self.style.layout("Custom.Treeview.Heading", [
            ("Treeheading.cell", {"sticky": "nswe"}),
            ("Treeheading.border", {"sticky": "nswe", "children": [
                ("Treeheading.padding", {"sticky": "nswe", "children": [
                    ("Treeheading.image", {"side": "right", "sticky": ""}),
                    ("Treeheading.text", {"sticky": "we"})
                ]})
            ]})
        ])
        
        for col, config in columns_config.items():
            # Criar estilo específico para cada coluna
            style_name = f"Heading.{col}"
            self.style.configure(
                style_name,
                background=config["background"],
                foreground=config["foreground"],
                font=("Segoe UI", 13, "bold"),
                relief="flat",
                borderwidth=0,
                padding=15
            )
            
            # Configurar a coluna e o cabeçalho
            self.results_tree.column(
                col,
                width=config["width"],
                anchor=config["anchor"],
                stretch=config["stretch"]
            )
            
            # Aplicar o cabeçalho com o estilo
            self.results_tree.heading(
                col, 
                text=col,
                command=lambda c=col: self.treeview_sort_column(c, False)
            )

        # Configurar scrollbar
        style_scroll = ttk.Style()
        style_scroll.configure(
            "Custom.Vertical.TScrollbar", 
            background="#2C3E50",
            troughcolor="#F5F5F5",
            width=10,
            arrowsize=10,
            relief="flat",
            borderwidth=0
        )

        scrollbar_y = ttk.Scrollbar(
            self.tree_frame,
            orient=tk.VERTICAL,
            command=self.results_tree.yview,
            style="Custom.Vertical.TScrollbar"
        )
        
        self.results_tree.configure(yscrollcommand=scrollbar_y.set)

        # Layout final
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.results_tree.pack(fill=tk.BOTH, expand=True)

        # Bind eventos
        self.results_tree.bind("<Double-1>", self.on_tree_double_click)
        self.results_tree.bind("<Button-3>", self.show_context_menu)
        self.results_tree.bind("<Return>", lambda e: self.on_tree_double_click(None))
        self.results_tree.bind("<space>", lambda e: self.toggle_row_selection())

        # Botão para exportar resultados
        self.export_button = ttk.Button(self.results_search_frame, text="Exportar para XLSX", command=self.export_to_xlsx, style="TButton")
        self.export_button.grid(row=1, column=1, padx=(5, 0), pady=(5, 0))


        # Ajustando a largura da coluna para o search_frame
        self.search_frame.columnconfigure(0, weight=1)
        self.search_frame.columnconfigure(1, weight=0)
    
    def toggle_entry(self, var, entry):
        if var.get() == 1:
            entry.config(state=tk.NORMAL)
        else:
            entry.config(state=tk.DISABLED)
            entry.delete(0, tk.END)

    def clear_checkboxes(self):
        for formula in self.formulas.values():
            formula['var'].set(0)
            formula['weight_entry'].delete(0, tk.END)
            formula['weight_entry'].config(state=tk.DISABLED)
        self.show_all_formulas()  # Exibir todas as fórmulas após limpar as checkboxes

    def open_html_file(self):
        webbrowser.open('index.html')


    def search_formula(self):
        formula_number = self.search_entry.get().strip()
        if formula_number:
            try:
                formula_id = int(formula_number)
                self.display_formula(formula_id)
            except ValueError:
                messagebox.showerror("Erro", "Número de fórmula inválido.")
        else:
            self.display_all_formulas()

    def display_formula(self, formula_id):
        for formula_frame in self.form.winfo_children():
            formula_frame.pack_forget()
        if formula_id in self.formulas:
            self.formulas[formula_id]['checkbox'].master.pack(fill=tk.X, pady=5)

    def display_all_formulas(self):
        for formula_frame in self.form.winfo_children():
            formula_frame.pack(fill=tk.X, pady=5)

    def show_all_formulas(self):
        self.filtered_form_section.pack_forget()  # Esconder a seção de fórmulas filtradas
        self.form_section.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)  # Mostrar a seção de todas as fórmulas
        self.display_all_formulas()

    def calculate(self):
        self.results_tree.delete(*self.results_tree.get_children())
        self.calculated_results = []
        for formula_id, data in self.formulas.items():
            if data['var'].get() == 1:
                try:
                    weight = float(data['weight_entry'].get().strip())
                    for formula in formula_data[formula_id]:
                        description = formula['description']
                        fixed_amount = formula['fixed_amount']
                        type_ = formula['type']
                        obs = formula['obs']
                        calculated_amount = "{:.2f}".format(fixed_amount * weight)
                        self.results_tree.insert("", tk.END, values=(
                            str(formula_id),
                            description,
                            calculated_amount,
                            type_,
                            obs
                        ))
                        self.calculated_results.append((
                            formula_id,
                            description,
                            calculated_amount,
                            type_,
                            obs
                        ))
                except ValueError:
                    messagebox.showerror("Erro", f"Peso inválido para a fórmula {formula_id}.")
                    return
        
        self.original_results = self.calculated_results.copy()
        self.update_row_colors()
        self.atualizar_soma_kg()
        self.adjust_column_width()  # Chamada da nova função após inserir os dados

    def filter_results(self):
        query = self.results_search_entry.get().strip().lower()
        if query:
            filtered_results = [
                result for result in self.original_results
                if query in result[1].lower()
            ]
            self.results_tree.delete(*self.results_tree.get_children())
            for result in filtered_results:
                self.results_tree.insert("", tk.END, values=result)
        else:
            messagebox.showwarning("Atenção", "Digite um termo de pesquisa.")
        self.filtro = "filtered"
        self.atualizar_soma_kg()

    def clear_filter(self):
        self.results_tree.delete(*self.results_tree.get_children())
        for result in self.original_results:
            self.results_tree.insert("", tk.END, values=result)
        self.filtro = None
        self.atualizar_soma_kg()


    def atualizar_soma_kg(self):
        soma_kg = 0
        # Obter todos os itens da treeview, independentemente do filtro
        itens = self.results_tree.get_children()
        for item in itens:
            try:
                kg = float(self.results_tree.item(item, "values")[2])
                soma_kg += kg
            except ValueError:
                continue
                
        # Configurar o label do texto com fonte maior
        self.soma_kg_label.config(
            text="Soma dos Kg: ",
            font=("Segoe UI", 16, "bold")  # Aumentado de 12 para 16
        )
        
        # Configurar ou criar o label do valor
        if not hasattr(self, 'valor_kg_label'):
            self.valor_kg_label = ttk.Label(
                self.results_search_frame,
                font=("Segoe UI", 16, "bold"),  # Aumentado de 12 para 16
                foreground="#FF0000"
            )
            self.valor_kg_label.grid(row=2, column=1, padx=(0, 5), pady=(5, 0), sticky='w')  # Adicionado sticky='w'
            
        # Atualizar o valor
        self.valor_kg_label.config(text=f"{soma_kg:.2f}")
        
        # Ajustar posicionamento
        self.soma_kg_label.grid(row=2, column=0, padx=(5, 0), pady=(5, 0), sticky='e')  # Adicionado sticky='e'

    
    def export_to_xlsx(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Resultados"

        # Adicionar cabeçalhos
        headers = ["Fórmula", "Descrição", "Quantidade Fixa", "Tipo"]
        sheet.append(headers)

        # Formatando cabeçalhos com fundo laranja e letra branca
        bold_font = Font(bold=True, color="FFFFFF")  # Texto em branco
        fill = PatternFill(start_color="FF6B2B", end_color="FF6B2B", fill_type="solid")  # Laranja
        center_alignment = Alignment(horizontal='center', vertical='center')  # Adiciona alinhamento central
        
        for cell in sheet["1:1"]:
            cell.font = bold_font
            cell.fill = fill
            cell.alignment = center_alignment  # Aplica o alinhamento

        # Adicionar dados
        for row_id in self.results_tree.get_children():
            row = self.results_tree.item(row_id)['values']
            sheet.append(row)

        # Ajustar largura das colunas
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        # Diálogo para escolher o local e nome do arquivo
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Salvar Arquivo como"
        )

        if file_path:  # Verifica se o usuário selecionou um caminho
            try:
                workbook.save(file_path)
                messagebox.showinfo("Exportação Completa", f"Os resultados foram exportados com sucesso para '{file_path}'.")
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível salvar o arquivo: {e}")
        else:
            messagebox.showwarning("Aviso", "Nenhum arquivo foi selecionado para salvar.")

    def treeview_sort_column(self, col, reverse):
        """Função para ordenar o treeview ao clicar nas colunas"""
        l = [(self.results_tree.set(k, col), k) for k in self.results_tree.get_children("")]
        try:
            l.sort(key=lambda t: float(t[0]), reverse=reverse)
        except ValueError:
            l.sort(reverse=reverse)

        for index, (val, k) in enumerate(l):
            self.results_tree.move(k, "", index)
            
        self.results_tree.heading(col, command=lambda: self.treeview_sort_column(col, not reverse))
        
        # Atualizar cores das linhas após ordenação
        self.update_row_colors()

    def update_row_colors(self):
        """Atualiza as cores das linhas alternadas"""
        for i, item in enumerate(self.results_tree.get_children()):
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            self.results_tree.item(item, tags=(tag,))

    def on_tree_double_click(self, event):
        """Manipula duplo clique em um item"""
        item = self.results_tree.selection()[0]
        values = self.results_tree.item(item)['values']
        self.show_details_dialog(values)

    def show_details_dialog(self, values):
        """Mostra diálogo com detalhes do item"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Detalhes da Fórmula")
        dialog.geometry("400x300")
        
        # Adicionar detalhes
        ttk.Label(dialog, text=f"Fórmula: {values[0]}", font=("Segoe UI", 12, "bold")).pack(pady=5)
        ttk.Label(dialog, text=f"Descrição: {values[1]}", wraplength=350).pack(pady=5)
        ttk.Label(dialog, text=f"Quantidade: {values[2]} kg").pack(pady=5)
        ttk.Label(dialog, text=f"Tipo: {values[3]}").pack(pady=5)
        if len(values) > 4:
            ttk.Label(dialog, text=f"Observação: {values[4]}", wraplength=350).pack(pady=5)

    def show_context_menu(self, event):
        """Mostra menu de contexto ao clicar com botão direito"""
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Copiar", command=self.copy_selected)
        menu.add_command(label="Exportar Selecionados", command=self.export_selected)
        menu.add_separator()
        menu.add_command(label="Detalhes", command=lambda: self.show_details_dialog(
            self.results_tree.item(self.results_tree.selection()[0])['values']
        ))
        
        menu.post(event.x_root, event.y_root)

    def copy_selected(self):
        """Copia item selecionado para clipboard"""
        selected = self.results_tree.selection()
        if selected:
            values = self.results_tree.item(selected[0])['values']
            self.root.clipboard_clear()
            self.root.clipboard_append("\t".join(map(str, values)))

    def export_selected(self):
        """Exporta apenas os itens selecionados"""
        selected = self.results_tree.selection()
        if selected:
            self.export_to_xlsx(selected_only=selected)

    def on_mousewheel(self, event):
        """Manipula o evento de scroll do mouse"""
        if self.form_canvas.winfo_exists():  # Verifica se o widget ainda existe
            self.form_canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def on_canvas_configure(self, event):
        """Ajusta a largura do frame interno quando o canvas é redimensionado"""
        if self.form_canvas.winfo_exists():  # Verifica se o widget ainda existe
            # Atualiza a largura da janela interna para corresponder à largura do canvas
            self.form_canvas.itemconfig(self.form_canvas.find_withtag("all")[0], width=event.width)

    def _on_frame_configure(self, event=None):
        """Atualiza o scrollregion do canvas quando o tamanho do frame interno muda"""
        if self.form_canvas.winfo_exists():  # Verifica se o widget ainda existe
            self.form_canvas.configure(scrollregion=self.form_canvas.bbox("all"))

    def bind_scroll(self, event):
        """Ativa os eventos de scroll quando o mouse entra no frame"""
        self.form_canvas.bind_all("<MouseWheel>", self.on_mousewheel)
        self.form_canvas.bind_all("<Up>", lambda e: self.form_canvas.yview_scroll(-1, "units"))
        self.form_canvas.bind_all("<Down>", lambda e: self.form_canvas.yview_scroll(1, "units"))
        self.form_canvas.bind_all("<Prior>", lambda e: self.form_canvas.yview_scroll(-1, "pages"))
        self.form_canvas.bind_all("<Next>", lambda e: self.form_canvas.yview_scroll(1, "pages"))

    def unbind_scroll(self, event):
        """Desativa os eventos de scroll quando o mouse sai do frame"""
        self.form_canvas.unbind_all("<MouseWheel>")
        self.form_canvas.unbind_all("<Up>")
        self.form_canvas.unbind_all("<Down>")
        self.form_canvas.unbind_all("<Prior>")
        self.form_canvas.unbind_all("<Next>")

    def insert_treeview_item(self, values):
        """Insere item no treeview com formatação melhorada"""
        item_id = self.results_tree.insert("", tk.END, values=values)
        self.update_row_colors()
        return item_id

    def toggle_row_selection(self):
        """Alterna a seleção da linha atual"""
        selection = self.results_tree.selection()
        if selection:
            self.results_tree.selection_remove(selection[0])
        else:
            children = self.results_tree.get_children()
            if children:
                self.results_tree.selection_set(children[0])

    def show_historico(self):
        """Abre a janela de histórico"""
        try:
            # Preparar os dados para o histórico
            formula = self.search_entry.get()
            programacao = self.programacao_entry.get()
            if not programacao:
                programacao = f"Fórmula {formula}"
            
            # Criar instância do histórico
            historico = HistoricoFrame(self.root)
            
            # Se houver resultados calculados, adicionar ao histórico
            if self.calculated_results:
                historico.add_record(
                    data_registro=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    programacao=programacao,
                    observacao="",  # Você pode adicionar um campo para observações se desejar
                    resultados=self.calculated_results
                )
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir histórico: {str(e)}")

    def open_documentation(self):
        """Abre o arquivo de documentação do projeto"""
        try:
            # Determinar o caminho base do aplicativo
            if getattr(sys, 'frozen', False):
                # Se for executável empacotado
                base_path = sys._MEIPASS
            else:
                # Se for script Python
                base_path = os.path.dirname(os.path.abspath(__file__))
            
            doc_path = os.path.join(base_path, 'Doc_completa.docx')
            
            if os.path.exists(doc_path):
                if os.name == 'nt':  # Windows
                    os.startfile(doc_path)
                else:  # Linux/Mac
                    if sys.platform == 'darwin':  # Mac
                        Popen(['open', doc_path])
                    else:  # Linux
                        Popen(['xdg-open', doc_path])
            else:
                # Tentar encontrar no diretório do executável
                exe_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(__file__)
                doc_path = os.path.join(exe_dir, 'Doc_completa.docx')
                
                if os.path.exists(doc_path):
                    if os.name == 'nt':  # Windows
                        os.startfile(doc_path)
                    else:  # Linux/Mac
                        if sys.platform == 'darwin':  # Mac
                            Popen(['open', doc_path])
                        else:  # Linux
                            Popen(['xdg-open', doc_path])
                else:
                    messagebox.showerror(
                        "Erro",
                        f"Arquivo de documentação não encontrado!\n\n"
                        f"Caminhos procurados:\n"
                        f"1. {os.path.join(base_path, 'Doc_completa.docx')}\n"
                        f"2. {doc_path}\n\n"
                        "Verifique se o arquivo 'Doc_completa.docx' está na mesma pasta do programa."
                    )
        except Exception as e:
            messagebox.showerror(
                "Erro", 
                f"Erro ao abrir documentação:\n\n{str(e)}\n\n"
                "Verifique se o arquivo 'Doc_completa.docx' está na mesma pasta do programa e "
                "se você tem um programa padrão configurado para abrir arquivos .docx"
            )

if __name__ == "__main__":
    root = tk.Tk()
    app = CalculatorApp(root)
    root.mainloop()

