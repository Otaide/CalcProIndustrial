import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class HistoricoFrame:
    def __init__(self, parent):
        self.window = tk.Toplevel(parent)
        self.window.title("Histórico de Fórmulas")
        self.window.geometry("1200x800")
        
        # Garantir que o banco de dados seja criado no diretório correto
        self.db_path = os.path.join(os.path.dirname(__file__), 'historico.db')
        
        # Criar banco de dados
        self.create_database()
        
        # Configurar estilo
        self.style = ttk.Style()
        self.configure_styles()
        
        self.create_widgets()
        self.load_history()

    def configure_styles(self):
        # Cores do tema - Esquema simplificado
        self.colors = {
            'primary': '#2196F3',      # Azul principal
            'background': '#F5F5F5',   # Cinza muito claro para fundo
            'text': '#212121',         # Cinza escuro para texto
            'white': '#FFFFFF',        # Branco
            'hover': '#1976D2',        # Azul mais escuro para hover
            'edit': '#4CAF50',         # Verde para edição
            'delete': '#F44336',       # Vermelho para exclusão
            'search': '#2196F3',       # Azul para busca
            'header': '#E0E0E0'        # Cinza claro para cabeçalhos
        }

        # Estilo da janela principal
        self.window.configure(bg=self.colors['background'])
        
        # Estilo para frames
        self.style.configure(
            "App.TFrame",
            background=self.colors['background']
        )
        
        # Estilo para LabelFrames
        self.style.configure(
            "App.TLabelframe",
            background=self.colors['background']
        )
        self.style.configure(
            "App.TLabelframe.Label",
            background=self.colors['background'],
            foreground=self.colors['text'],
            font=("Segoe UI", 11, "bold")
        )

        # Estilo para os Treeviews
        self.style.configure(
            "Historico.Treeview",
            background=self.colors['white'],
            foreground=self.colors['text'],
            rowheight=30,
            fieldbackground=self.colors['white']
        )
        
        # Estilo para os cabeçalhos do Treeview
        self.style.configure(
            "Historico.Treeview.Heading",
            background=self.colors['header'],
            foreground=self.colors['text'],
            font=("Segoe UI", 10, "bold")
        )
        self.style.map("Historico.Treeview.Heading",
            background=[("active", self.colors['header'])]
        )

        # Configurar seleção do Treeview
        self.style.map("Historico.Treeview",
            background=[("selected", self.colors['primary'])],
            foreground=[("selected", self.colors['white'])]
        )

        # Estilo para botões padrão
        self.style.configure(
            "App.TButton",
            background=self.colors['primary'],
            foreground=self.colors['text'],
            font=("Segoe UI", 10),
            padding=(10, 5)
        )
        self.style.map("App.TButton",
            background=[("active", self.colors['hover'])]
        )

        # Estilo para botão de pesquisa
        self.style.configure(
            "Search.TButton",
            background=self.colors['search'],
            foreground=self.colors['text'],
            padding=(10, 5),
            font=("Segoe UI", 10)
        )
        self.style.map("Search.TButton",
            background=[("active", self.colors['hover'])]
        )

        # Estilo para botão de exclusão
        self.style.configure(
            "Danger.TButton",
            background="#e0e0e0",        # Cinza claro
            foreground="#d32f2f",        # Vermelho escuro para o texto
            padding=(10, 5),
            font=("Segoe UI", 10)
        )
        self.style.map("Danger.TButton",
            background=[("active", "#d5d5d5")],  # Cinza um pouco mais escuro no hover
            foreground=[("active", "#b71c1c")]   # Vermelho mais escuro no hover
        )

        # Estilo para botão de edição
        self.style.configure(
            "Edit.TButton",
            background="#e0e0e0",        # Cinza claro
            foreground="#1976d2",        # Azul escuro para o texto
            padding=(10, 5),
            font=("Segoe UI", 10)
        )
        self.style.map("Edit.TButton",
            background=[("active", "#d5d5d5")],  # Cinza um pouco mais escuro no hover
            foreground=[("active", "#1565c0")]   # Azul mais escuro no hover
        )

        # Estilo para labels
        self.style.configure(
            "App.TLabel",
            background=self.colors['background'],
            foreground=self.colors['text'],
            font=("Segoe UI", 10)
        )

        # Estilo para entries
        self.style.configure(
            "App.TEntry",
            fieldbackground=self.colors['white'],
            foreground=self.colors['text']
        )

        # Estilo para botão de exportação
        self.style.configure(
            "Export.TButton",
            background="#e0e0e0",        # Cinza claro
            foreground="#2e7d32",        # Verde escuro para o texto
            padding=(10, 5),
            font=("Segoe UI", 10)
        )
        self.style.map("Export.TButton",
            background=[("active", "#d5d5d5")],  # Cinza mais escuro no hover
            foreground=[("active", "#1b5e20")]   # Verde mais escuro no hover
        )

        # Estilo para botão de retorno
        self.style.configure(
            "Return.TButton",
            background="#e0e0e0",        # Cinza claro
            foreground="#1976d2",        # Azul para o texto
            padding=(10, 5),
            font=("Segoe UI", 10)
        )
        self.style.map("Return.TButton",
            background=[("active", "#d5d5d5")],  # Cinza mais escuro no hover
            foreground=[("active", "#1565c0")]   # Azul mais escuro no hover
        )

        # Estilo para botão de exclusão total
        self.style.configure(
            "DeleteAll.TButton",
            background="#e0e0e0",        # Cinza claro
            foreground="#c62828",        # Vermelho mais escuro para o texto
            padding=(10, 5),
            font=("Segoe UI", 10, "bold")
        )
        self.style.map("DeleteAll.TButton",
            background=[("active", "#d5d5d5")],  # Cinza mais escuro no hover
            foreground=[("active", "#b71c1c")]   # Vermelho ainda mais escuro no hover
        )

    def create_widgets(self):
        # Frame principal com novo estilo
        main_frame = ttk.Frame(self.window, padding=10, style="App.TFrame")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Frame superior para informações gerais
        info_frame = ttk.LabelFrame(
            main_frame, 
            text="Informações Gerais", 
            padding=10,
            style="App.TLabelframe"
        )
        info_frame.pack(fill=tk.X, pady=(0, 10))

        # Frame para pesquisa
        search_frame = ttk.Frame(info_frame, style="App.TFrame")
        search_frame.pack(fill=tk.X, pady=(0, 5))

        # Campo de pesquisa com novo estilo
        ttk.Label(
            search_frame, 
            text="Pesquisar:", 
            style="App.TLabel"
        ).pack(side=tk.LEFT, padx=(0, 5))
        
        self.search_entry = ttk.Entry(
            search_frame, 
            width=40,
            style="App.TEntry"
        )
        self.search_entry.pack(side=tk.LEFT, padx=5)
        
        # Botões com novos estilos
        ttk.Button(
            search_frame,
            text="Buscar",
            command=self.search_records,
            style="Search.TButton"
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            search_frame,
            text="Limpar Busca",
            command=self.clear_search,
            style="App.TButton"
        ).pack(side=tk.LEFT, padx=5)

        # Frame para botões de ação
        button_frame = ttk.Frame(info_frame, style="App.TFrame")
        button_frame.pack(fill=tk.X, pady=(5, 0))

        # Botão Calculadora de Fórmulas (à direita)
        ttk.Button(
            button_frame,
            text="Calculadora de Fórmulas",
            command=self.voltar_calculadora,
            style="Return.TButton"
        ).pack(side=tk.RIGHT, padx=5)

        # Botões à esquerda
        ttk.Button(
            button_frame,
            text="Editar",
            command=self.edit_record,
            style="Edit.TButton"
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text="Excluir",
            command=self.delete_record,
            style="Danger.TButton"
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text="Exportar",
            command=self.export_results,
            style="Export.TButton"
        ).pack(side=tk.LEFT, padx=5)

        # Botão Excluir Todos movido para depois do Exportar
        ttk.Button(
            button_frame,
            text="Excluir Todos",
            command=self.delete_all_records,
            style="DeleteAll.TButton"
        ).pack(side=tk.LEFT, padx=5)

        # Treeview para informações gerais
        self.info_tree = ttk.Treeview(
            info_frame,
            columns=("Data", "Programação", "Observação"),
            show="headings",
            style="Historico.Treeview",
            height=5
        )

        # Configurar colunas do info_tree
        info_columns = {
            "Data": {"width": 150, "anchor": "center"},
            "Programação": {"width": 300, "anchor": "w"},
            "Observação": {"width": 500, "anchor": "w"}
        }

        for col, config in info_columns.items():
            self.info_tree.heading(col, text=col)
            self.info_tree.column(col, width=config["width"], anchor=config["anchor"])

        # Scrollbar para info_tree
        info_scrollbar = ttk.Scrollbar(info_frame, orient="vertical", command=self.info_tree.yview)
        self.info_tree.configure(yscrollcommand=info_scrollbar.set)

        # Layout do info_tree e scrollbar
        self.info_tree.pack(side="left", fill="both", expand=True)
        info_scrollbar.pack(side="right", fill="y")

        # Frame inferior para resultados detalhados
        results_frame = ttk.LabelFrame(main_frame, text="Resultados Detalhados", padding=10)
        results_frame.pack(fill=tk.BOTH, expand=True)

        # Frame para pesquisa nos resultados detalhados
        results_search_frame = ttk.Frame(results_frame, style="App.TFrame")
        results_search_frame.pack(fill=tk.X, pady=(0, 10))

        # Frame para pesquisa e soma
        search_sum_frame = ttk.Frame(results_search_frame, style="App.TFrame")
        search_sum_frame.pack(fill=tk.X)

        # Lado esquerdo - Pesquisa
        search_left_frame = ttk.Frame(search_sum_frame, style="App.TFrame")
        search_left_frame.pack(side=tk.LEFT)

        ttk.Label(
            search_left_frame,
            text="Filtrar por Descrição:",
            font=("Segoe UI", 10),
            style="App.TLabel"
        ).pack(side=tk.LEFT, padx=(0, 5))

        self.results_search_entry = ttk.Entry(
            search_left_frame,
            width=40,
            font=("Segoe UI", 10),
            style="App.TEntry"
        )
        self.results_search_entry.pack(side=tk.LEFT, padx=5)

        # Lado direito - Soma
        sum_frame = ttk.Frame(search_sum_frame, style="App.TFrame")
        sum_frame.pack(side=tk.RIGHT, padx=10)

        self.sum_label = ttk.Label(
            sum_frame,
            text="Soma dos Kg: 0.00",
            font=("Segoe UI", 14, "bold"),
            foreground="#1976d2",
            style="App.TLabel"
        )
        self.sum_label.pack(side=tk.RIGHT)

        # Binds para a pesquisa
        self.results_search_entry.bind('<Return>', lambda e: self.filter_results_details())
        self.results_search_entry.bind('<Escape>', lambda e: self.clear_results_filter())
        
        # Treeview para resultados
        self.results_tree = ttk.Treeview(
            results_frame,
            columns=("Fórmula", "Descrição", "Kg", "Tipo", "Observação"),
            show="headings",
            style="Historico.Treeview"
        )

        # Configurar colunas do results_tree
        results_columns = {
            "Fórmula": {"width": 100, "anchor": "center"},
            "Descrição": {"width": 400, "anchor": "w"},
            "Kg": {"width": 100, "anchor": "center"},
            "Tipo": {"width": 150, "anchor": "center"},
            "Observação": {"width": 300, "anchor": "w"}
        }

        for col, config in results_columns.items():
            self.results_tree.heading(col, text=col)
            self.results_tree.column(col, width=config["width"], anchor=config["anchor"])
        
        # Scrollbar para results_tree
        results_scrollbar = ttk.Scrollbar(results_frame, orient="vertical", command=self.results_tree.yview)
        self.results_tree.configure(yscrollcommand=results_scrollbar.set)
        
        # Layout do results_tree
        self.results_tree.pack(side="left", fill="both", expand=True)
        results_scrollbar.pack(side="right", fill="y")

        # Bind para seleção no info_tree
        self.info_tree.bind('<<TreeviewSelect>>', self.on_info_select)

    def load_history(self):
        try:
            conn = sqlite3.connect(self.db_path)
            c = conn.cursor()
            c.execute("SELECT data, programacao, obs FROM historico ORDER BY data DESC")
            rows = c.fetchall()
            
            for item in self.info_tree.get_children():
                self.info_tree.delete(item)
                
            for row in rows:
                self.info_tree.insert("", "end", values=row)
                
        except sqlite3.Error as e:
            print(f"Erro ao carregar histórico: {e}")
        finally:
            conn.close()

    def on_info_select(self, event):
        selection = self.info_tree.selection()
        if selection:
            item = self.info_tree.item(selection[0])
            data = item['values'][0]
            
            for item in self.results_tree.get_children():
                self.results_tree.delete(item)
            
            try:
                conn = sqlite3.connect(self.db_path)
                c = conn.cursor()
                c.execute("""
                    SELECT formula, descricao, kg, tipo, obs 
                    FROM resultados 
                    WHERE data = ?
                    ORDER BY formula
                """, (data,))
                rows = c.fetchall()
                for row in rows:
                    self.results_tree.insert("", "end", values=row)
            except sqlite3.Error as e:
                print(f"Erro ao carregar resultados: {e}")
            finally:
                conn.close()

    def add_record(self, data_registro, programacao, observacao, resultados):
        try:
            conn = sqlite3.connect(self.db_path)
            c = conn.cursor()
            
            c.execute("INSERT INTO historico (data, programacao, obs) VALUES (?, ?, ?)",
                     (data_registro, programacao, observacao))
            
            for resultado in resultados:
                c.execute("""
                    INSERT INTO resultados (data, formula, descricao, kg, tipo, obs) 
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (data_registro, str(resultado[0]), str(resultado[1]), 
                     float(resultado[2]), str(resultado[3]), str(resultado[4])))
            
            conn.commit()
            self.load_history()
            
        except sqlite3.Error as e:
            print(f"Erro ao adicionar registro: {e}")
        finally:
            conn.close()

    def create_database(self):
        try:
            # Se o banco de dados existir e estiver corrompido, deletar
            if os.path.exists(self.db_path):
                try:
                    conn = sqlite3.connect(self.db_path)
                    conn.close()
                except sqlite3.Error:
                    os.remove(self.db_path)

            conn = sqlite3.connect(self.db_path)
            c = conn.cursor()
            
            c.execute('''CREATE TABLE IF NOT EXISTS historico
                        (data TEXT, programacao TEXT, obs TEXT)''')
            
            c.execute('''CREATE TABLE IF NOT EXISTS resultados
                        (data TEXT, formula TEXT, descricao TEXT, 
                         kg REAL, tipo TEXT, obs TEXT)''')
            
            conn.commit()
            
        except sqlite3.Error as e:
            print(f"Erro ao criar banco de dados: {e}")
        finally:
            conn.close() 
            
    def search_records(self):
        """Pesquisar registros"""
        search_term = self.search_entry.get().strip().lower()
        if not search_term:
            self.load_history()
            return

        try:
            conn = sqlite3.connect(self.db_path)
            c = conn.cursor()
            c.execute("""
                SELECT data, programacao, obs 
                FROM historico 
                WHERE LOWER(programacao) LIKE ? OR LOWER(obs) LIKE ?
                ORDER BY data DESC
            """, (f'%{search_term}%', f'%{search_term}%'))
            
            rows = c.fetchall()
            
            # Limpar a árvore
            for item in self.info_tree.get_children():
                self.info_tree.delete(item)
                
            # Inserir resultados encontrados
            for row in rows:
                self.info_tree.insert("", "end", values=row)
                
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao pesquisar registros: {e}")
        finally:
            conn.close()

    def clear_search(self):
        """Limpar pesquisa e mostrar todos os registros"""
        self.search_entry.delete(0, tk.END)
        self.load_history()

    def edit_record(self):
        """Editar registro selecionado"""
        selection = self.info_tree.selection()
        if not selection:
            messagebox.showwarning("Aviso", "Selecione um registro para editar")
            return

        item = self.info_tree.item(selection[0])
        values = item['values']

        # Criar janela de edição com tamanho maior
        edit_window = tk.Toplevel(self.window)
        edit_window.title("Editar Registro")
        edit_window.geometry("600x400")  # Aumentado de 500x300 para 600x400
        edit_window.configure(bg='#f0f0f0')
        
        # Tornar a janela modal
        edit_window.transient(self.window)
        edit_window.grab_set()
        
        # Frame principal com padding maior
        main_frame = ttk.Frame(edit_window, style="App.TFrame")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)  # Aumentado padding

        # Título com mais espaço
        title_label = ttk.Label(
            main_frame,
            text="Editar Registro",
            font=("Segoe UI", 16, "bold"),  # Fonte maior
            style="App.TLabel"
        )
        title_label.pack(pady=(0, 30))  # Mais espaço após o título

        # Frame para os campos com mais espaço
        fields_frame = ttk.Frame(main_frame, style="App.TFrame")
        fields_frame.pack(fill=tk.X, padx=20)  # Aumentado padding

        # Campo Programação
        prog_frame = ttk.Frame(fields_frame, style="App.TFrame")
        prog_frame.pack(fill=tk.X, pady=10)  # Aumentado spacing
        
        ttk.Label(
            prog_frame,
            text="Programação:",
            font=("Segoe UI", 11, "bold"),  # Fonte maior
            style="App.TLabel"
        ).pack(anchor=tk.W)
        
        prog_entry = ttk.Entry(
            prog_frame,
            width=60,  # Aumentado largura
            font=("Segoe UI", 11),  # Fonte maior
            style="App.TEntry"
        )
        prog_entry.pack(fill=tk.X, pady=(8, 0))  # Mais espaço
        prog_entry.insert(0, values[1])

        # Campo Observação
        obs_frame = ttk.Frame(fields_frame, style="App.TFrame")
        obs_frame.pack(fill=tk.X, pady=20)  # Mais espaço entre campos
        
        ttk.Label(
            obs_frame,
            text="Observação:",
            font=("Segoe UI", 11, "bold"),  # Fonte maior
            style="App.TLabel"
        ).pack(anchor=tk.W)
        
        obs_entry = ttk.Entry(
            obs_frame,
            width=60,  # Aumentado largura
            font=("Segoe UI", 11),  # Fonte maior
            style="App.TEntry"
        )
        obs_entry.pack(fill=tk.X, pady=(8, 0))  # Mais espaço
        obs_entry.insert(0, values[2])

        # Frame para botões com mais espaço
        button_frame = ttk.Frame(main_frame, style="App.TFrame")
        button_frame.pack(pady=30)  # Mais espaço antes dos botões

        def save_changes():
            try:
                conn = sqlite3.connect(self.db_path)
                c = conn.cursor()
                c.execute("""
                    UPDATE historico 
                    SET programacao = ?, obs = ?
                    WHERE data = ?
                """, (prog_entry.get(), obs_entry.get(), values[0]))
                conn.commit()
                conn.close()
                
                self.load_history()
                edit_window.destroy()
                messagebox.showinfo("Sucesso", "Registro atualizado com sucesso!")
            except sqlite3.Error as e:
                messagebox.showerror("Erro", f"Erro ao atualizar registro: {e}")

        # Estilo para o botão salvar
        self.style.configure(
            "Save.TButton",
            background="#e0e0e0",        # Cinza claro
            foreground="#1976d2",        # Azul para o texto
            padding=(15, 8),
            font=("Segoe UI", 10)
        )
        self.style.map("Save.TButton",
            background=[("active", "#d5d5d5")],  # Cinza mais escuro no hover
            foreground=[("active", "#1565c0")]   # Azul mais escuro no hover
        )

        # Botões
        ttk.Button(
            button_frame,
            text="Cancelar",
            command=edit_window.destroy,
            style="App.TButton",
            width=15
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text="Salvar Alterações",
            command=save_changes,
            style="Save.TButton",
            width=15
        ).pack(side=tk.LEFT, padx=5)

        # Centralizar a janela
        edit_window.update_idletasks()
        width = edit_window.winfo_width()
        height = edit_window.winfo_height()
        x = (edit_window.winfo_screenwidth() // 2) - (width // 2)
        y = (edit_window.winfo_screenheight() // 2) - (height // 2)
        edit_window.geometry(f'{width}x{height}+{x}+{y}')

    def delete_record(self):
        """Excluir registro selecionado"""
        selection = self.info_tree.selection()
        if not selection:
            messagebox.showwarning("Aviso", "Selecione um registro para excluir")
            return

        if not messagebox.askyesno("Confirmar", "Tem certeza que deseja excluir este registro?"):
            return

        item = self.info_tree.item(selection[0])
        data = item['values'][0]

        try:
            conn = sqlite3.connect(self.db_path)
            c = conn.cursor()
            
            # Excluir da tabela historico
            c.execute("DELETE FROM historico WHERE data = ?", (data,))
            
            # Excluir registros relacionados da tabela resultados
            c.execute("DELETE FROM resultados WHERE data = ?", (data,))
            
            conn.commit()
            conn.close()
            
            self.load_history()
            messagebox.showinfo("Sucesso", "Registro excluído com sucesso!")
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao excluir registro: {e}")

    def export_results(self):
        """Exportar resultados detalhados para Excel"""
        selection = self.info_tree.selection()
        if not selection:
            messagebox.showwarning("Aviso", "Selecione um registro para exportar")
            return

        item = self.info_tree.item(selection[0])
        info_values = item['values']
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Resultados Detalhados"

            # Cores
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            subheader_font = Font(color="000000", bold=True, size=11)
            data_font = Font(size=10)
            
            # Estilo para alinhamento
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')

            # Título do relatório
            ws.merge_cells('A1:E1')
            ws['A1'] = "Relatório de Resultados Detalhados"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = center_alignment
            ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            ws['A1'].font = Font(color="FFFFFF", size=14, bold=True)

            # Informações gerais
            ws['A3'] = "Data:"
            ws['B3'] = info_values[0]
            ws['A4'] = "Programação:"
            ws['B4'] = info_values[1]
            ws['A5'] = "Observação:"
            ws['B5'] = info_values[2]

            # Formatação das informações gerais
            for row in range(3, 6):
                ws[f'A{row}'].font = subheader_font
                ws[f'A{row}'].fill = subheader_fill
                ws[f'B{row}'].font = data_font
                ws[f'A{row}'].alignment = left_alignment
                ws[f'B{row}'].alignment = left_alignment

            # Linha em branco
            current_row = 7

            # Cabeçalhos das colunas
            headers = ["Fórmula", "Descrição", "Kg", "Tipo", "Observação"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

            # Adicionar dados
            for item in self.results_tree.get_children():
                current_row += 1
                valores = self.results_tree.item(item)['values']
                for col, valor in enumerate(valores, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = valor
                    cell.font = data_font
                    # Alinhar números à direita e texto à esquerda
                    if col == 3:  # Coluna Kg
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.alignment = left_alignment

                # Alternar cores das linhas
                if current_row % 2 == 0:
                    for col in range(1, 6):
                        ws.cell(row=current_row, column=col).fill = PatternFill(
                            start_color="F5F5F5", 
                            end_color="F5F5F5", 
                            fill_type="solid"
                        )

            # Ajustar largura das colunas
            column_widths = [15, 40, 12, 20, 30]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # Adicionar bordas
            thin_border = Side(border_style="thin", color="000000")
            border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
            
            for row in ws.iter_rows(min_row=7, max_row=current_row, min_col=1, max_col=5):
                for cell in row:
                    cell.border = border

            # Gerar nome do arquivo com timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"resultados_detalhados_{timestamp}.xlsx"

            # Abrir diálogo para escolher onde salvar
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=default_filename,
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Salvar Resultados Como"
            )
            
            if filename:
                wb.save(filename)
                messagebox.showinfo("Sucesso", f"Dados exportados com sucesso para:\n{filename}")
            else:
                messagebox.showinfo("Info", "Exportação cancelada pelo usuário")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar dados: {str(e)}") 

    def voltar_calculadora(self):
        """Fecha a janela do histórico e retorna para a calculadora"""
        self.window.destroy() 

    def filter_results_details(self):
        """Filtrar resultados detalhados baseado na descrição"""
        search_term = self.results_search_entry.get().strip().lower()
        if not search_term:
            return

        # Guardar todos os itens atuais
        all_items = []
        for item in self.results_tree.get_children():
            values = self.results_tree.item(item)['values']
            all_items.append(values)

        # Limpar a árvore
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)

        # Inserir apenas os itens filtrados e calcular soma
        soma_kg = 0.0
        for item in all_items:
            if search_term in str(item[1]).lower():  # item[1] é a coluna Descrição
                self.results_tree.insert("", "end", values=item)
                try:
                    soma_kg += float(item[2])  # item[2] é a coluna Kg
                except (ValueError, TypeError):
                    pass

        # Atualizar label com a soma
        self.sum_label.config(
            text=f"Soma dos Kg: {soma_kg:.2f}",
            font=("Segoe UI", 14, "bold")  # Mantendo a fonte maior
        )

    def clear_results_filter(self, event=None):
        """Limpar filtro e mostrar todos os resultados"""
        self.results_search_entry.delete(0, tk.END)
        
        # Recarregar os resultados do registro selecionado
        selection = self.info_tree.selection()
        if selection:
            self.on_info_select(None)  # Recarrega os resultados originais
        
        # Calcular soma total
        soma_kg = 0.0
        for item in self.results_tree.get_children():
            try:
                kg = float(self.results_tree.item(item)['values'][2])
                soma_kg += kg
            except (ValueError, TypeError):
                pass
        
        # Atualizar label com a soma total
        self.sum_label.config(
            text=f"Soma dos Kg: {soma_kg:.2f}",
            font=("Segoe UI", 14, "bold")  # Mantendo a fonte maior
        ) 

    def delete_all_records(self):
        """Excluir todos os registros do histórico"""
        if not messagebox.askyesno(
            "Confirmar Exclusão", 
            "ATENÇÃO: Isso excluirá TODOS os registros do histórico!\n\nTem certeza que deseja continuar?",
            icon='warning'
        ):
            return

        # Segunda confirmação para garantir
        if not messagebox.askyesno(
            "Confirmação Final",
            "Esta ação não pode ser desfeita!\n\nDeseja realmente excluir todos os registros?",
            icon='warning'
        ):
            return

        try:
            conn = sqlite3.connect(self.db_path)
            c = conn.cursor()
            
            # Excluir todos os registros das duas tabelas
            c.execute("DELETE FROM historico")
            c.execute("DELETE FROM resultados")
            
            conn.commit()
            conn.close()
            
            # Recarregar a visualização
            self.load_history()
            
            # Limpar a árvore de resultados
            self.results_tree.delete(*self.results_tree.get_children())
            
            messagebox.showinfo("Sucesso", "Todos os registros foram excluídos com sucesso!")
        
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao excluir registros: {e}") 